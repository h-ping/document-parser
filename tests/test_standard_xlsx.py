import json
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from document_parser.standard_xlsx import StandardXlsxParser


ROOT = Path(__file__).resolve().parents[1]


class StandardXlsxParserTests(unittest.TestCase):
    def test_xlsx_parser_reads_whitelist_sheets_and_builds_contract_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            xlsx_path = temp_path / "standard.xlsx"
            output_dir = temp_path / "out"
            _write_standard_xlsx_fixture(xlsx_path)

            result = StandardXlsxParser().parse(xlsx_path, debug_dir=output_dir)

            self.assertEqual(result.job["job_type"], "standard_xlsx_to_structured_json")
            self.assertEqual(result.metadata["output_contract_validation_report"]["status"], "pass")
            structure = json.loads((output_dir / "xlsx_workbook_structure.json").read_text(encoding="utf-8"))
            self.assertEqual(structure["read_sheets"], ["标签主文字", "企业信息", "营养表清单", "营养项目明细", "多内容物组合装", "其他标签文字"])
            self.assertIn("导出摘要", structure["ignored_sheets"])
            self.assertIn("使用说明", structure["ignored_sheets"])

            standard_items = result.metadata["standard_artifacts"]["standard_items"]
            comparison_entries = result.metadata["standard_artifacts"]["comparison_index"]["entries"]
            printable_text = "\n".join(item["text"] for item in standard_items)
            comparison_text = "\n".join(entry["text"] for entry in comparison_entries)
            for forbidden in ("MAIN_", "字段编码", "记录编号", "内容物编号", "不要输出备注", "不要读取"):
                self.assertNotIn(forbidden, printable_text)
                self.assertNotIn(forbidden, comparison_text)
            self.assertIn("10%", printable_text)
            self.assertNotIn("0.1", printable_text)

            tables = result.metadata["standard_artifacts"]["tables"]
            self.assertEqual([table["table_id"] for table in tables], ["nutrition_n1", "nutrition_n2"])
            self.assertEqual([column["name"] for column in tables[0]["columns"]], ["项目", "每 100 克", "营养素参考值%"])
            nutrition_texts = [item["text"] for item in standard_items if item["semantic_key"] == "product.nutrition_table"]
            self.assertTrue(any("底部备注文字" in text for text in nutrition_texts))

            content_groups = [group for group in result.metadata["standard_artifacts"]["field_groups"] if group["group_type"] == "content_item"]
            self.assertEqual(len(content_groups), 1)
            self.assertEqual(content_groups[0]["linked_table_ids"], [])
            self.assertTrue(any(ref.startswith("xlsx:标签主文字!") for evidence in result.evidence for ref in evidence.source_node_ids))
            preview_html = (output_dir / "result_preview.html").read_text(encoding="utf-8")
            self.assertIn("结构化标签内容预览", preview_html)
            self.assertIn("compare-shell", preview_html)
            self.assertIn("doc-pane", preview_html)
            self.assertIn("extract-card", preview_html)
            self.assertIn("group-card", preview_html)
            self.assertIn("table-extract-card", preview_html)
            self.assertIn("sheet-card", preview_html)
            self.assertNotIn("<th>字段编码</th>", preview_html)
            self.assertNotIn("<th>记录编号</th>", preview_html)
            self.assertNotIn("<th>营养表编号</th>", preview_html)
            self.assertNotIn("<th>内容物编号</th>", preview_html)
            self.assertNotIn("<th>备注</th>", preview_html)
            self.assertNotIn("不要输出备注", preview_html)
            self.assertNotIn("ignored-cell", preview_html)
            self.assertIn("<th>标签项目</th>", preview_html)
            self.assertIn("<th>客户填写原文</th>", preview_html)
            self.assertIn("<th>企业角色</th>", preview_html)
            self.assertIn("<th>底部备注/脚注原文</th>", preview_html)
            self.assertIn("<th>每 100 克</th>", preview_html)
            self.assertIn("委托方信息", preview_html)
            self.assertIn("广州酒家集团股份有限公司", preview_html)
            self.assertIn("广州市荔湾区", preview_html)
            self.assertIn("粽子 A 营养成分表", preview_html)
            self.assertIn("项目", preview_html)
            self.assertIn("每 100 克", preview_html)
            self.assertIn("营养素参考值%", preview_html)
            self.assertNotIn("每100g/mL含量", preview_html)
            self.assertIn("activateItem", preview_html)
            self.assertIn("xlsx:标签主文字!", preview_html)
            self.assertIn("bbox_status", preview_html)

    def test_manifest_xlsx_branch_does_not_require_ocr_or_llm_env(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            xlsx_path = temp_path / "standard.xlsx"
            manifest_path = temp_path / "manifest.json"
            output_dir = temp_path / "out"
            _write_standard_xlsx_fixture(xlsx_path)
            manifest_path.write_text(json.dumps({"input_xlsx": str(xlsx_path)}, ensure_ascii=False), encoding="utf-8")

            env = dict(os.environ)
            env.pop("GLM_OCR_API_KEY", None)
            env.pop("ZAI_API_KEY", None)
            env.pop("ZHIPUAI_API_KEY", None)
            env.pop("LLM_API_KEY", None)
            env.pop("LLM_BASE_URL", None)
            env.pop("LLM_MODEL", None)
            env["PYTHONPATH"] = os.pathsep.join(part for part in (str(ROOT / "src"), env.get("PYTHONPATH", "")) if part)
            completed = subprocess.run(
                [
                    sys.executable,
                    str(ROOT / "scripts" / "extract_structure.py"),
                    "--manifest",
                    str(manifest_path),
                    "--output-dir",
                    str(output_dir),
                ],
                cwd=ROOT,
                env=env,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                check=False,
            )

            self.assertEqual(completed.returncode, 0, completed.stderr)
            for artifact_name in (
                "standard_items.json",
                "comparison_index.json",
                "tables.json",
                "field_groups.json",
                "quality_report.json",
                "output_contract_validation_report.json",
                "xlsx_workbook_structure.json",
                "result_preview.html",
            ):
                self.assertTrue((output_dir / artifact_name).exists(), artifact_name)
            inventory = json.loads((output_dir / "00_inputs" / "file_inventory.json").read_text(encoding="utf-8"))
            self.assertIn("input_xlsx", {item["role"] for item in inventory["files"]})
            artifact_index = json.loads((output_dir / "artifacts" / "index.json").read_text(encoding="utf-8"))
            self.assertIn("result_preview.html", {item["path"] for item in artifact_index["artifacts"]})

    def test_nutrition_amount_header_accepts_unit_parentheses_alias(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            for amount_header in ("每100克(g)", "每份（65g）"):
                with self.subTest(amount_header=amount_header):
                    xlsx_path = temp_path / f"{amount_header}.xlsx"
                    _write_standard_xlsx_fixture(xlsx_path, amount_header=amount_header)

                    result = StandardXlsxParser().parse(xlsx_path)

                    tables = result.metadata["standard_artifacts"]["tables"]
                    self.assertEqual([column["name"] for column in tables[0]["columns"]], ["项目", amount_header, "营养素参考值%"])


def _write_standard_xlsx_fixture(path: Path, amount_header: str = "每 100 克") -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "导出摘要"
    summary.append(["不要读取"])
    instructions = workbook.create_sheet("使用说明")
    instructions.append(["不要读取"])

    main = workbook.create_sheet("标签主文字")
    main.append(["字段编码", "标签项目", "客户填写原文", "备注", "列1"])
    main.append(["MAIN_PRODUCT_NAME", "产品名称", "广式粽子", "不要输出备注", ""])
    main.append(["MAIN_NET_CONTENT", "净含量", "600 克", "", ""])
    main.append(["MAIN_INGREDIENTS", "配料", "糯米、猪肉", "", ""])
    main.append(["MAIN_STORAGE", "贮存条件", "常温保存", "", ""])
    main.append(["MAIN_SHELF_LIFE", "保质期", "9 个月", "", ""])
    main.append(["MAIN_STANDARD_CODE", "产品标准号", "GB/T 20977", "", ""])

    enterprise = workbook.create_sheet("企业信息")
    enterprise.append(["记录编号", "企业角色", "企业名称原文", "地址原文", "产地原文", "许可证/备案号", "联系方式", "邮政编码", "网站", "备注"])
    enterprise.append(["E1", "委托方", "广州酒家集团股份有限公司", "广州市荔湾区", "广东广州", "SC123", "400-0000", "510000", "www.example.com", "不要输出备注"])
    enterprise.append(["E2", "受委托方", "广州酒家利口福食品有限公司", "广州市番禺区", "广东广州", "SC456", "020-0000", "511400", "", ""])

    nutrition_list = workbook.create_sheet("营养表清单")
    nutrition_list.append(["营养表编号", "表名/适用范围", "底部备注/脚注原文", "备注"])
    nutrition_list.append(["N1", "粽子 A 营养成分表", "底部备注文字", "不要输出备注"])
    nutrition_list.append(["N2", "粽子 B 营养成分表", "", ""])

    nutrition_items = workbook.create_sheet("营养项目明细")
    nutrition_items.append(["营养表编号", "项目", amount_header, "营养素参考值%", "备注"])
    nutrition_items.append(["N1", "能量", "810 千焦", 0.1, "不要输出备注"])
    nutrition_items["D2"].number_format = "0%"
    nutrition_items.append(["N1", "蛋白质", "6.7 克", 0.11, ""])
    nutrition_items["D3"].number_format = "0%"
    nutrition_items.append(["N2", "能量", "900 千焦", 0.12, ""])
    nutrition_items["D4"].number_format = "0%"
    nutrition_items.append(["N3", "能量", "", "", "模板预置空行"])
    nutrition_items.append(["N3", "蛋白质", "", "", "模板预置空行"])

    content = workbook.create_sheet("多内容物组合装")
    content.append(["内容物编号", "内容物名称", "净含量/数量", "产品分类", "配料原文", "备注"])
    content.append(["C1", "粽子 A", "300 克", "真空包装熟制粽子", "糯米、猪肉", "不要输出备注"])

    other = workbook.create_sheet("其他标签文字")
    other.append(["记录编号", "标签文字类型", "客户填写原文", "备注"])
    other.append(["O1", "回收标识", "请按当地规定分类回收", "不要输出备注"])

    workbook.save(path)


if __name__ == "__main__":
    unittest.main()
