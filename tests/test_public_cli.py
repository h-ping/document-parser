import base64
import json
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from document_parser.consistency_cli import ConsistencyPipelineError, _ensure_ocr_token_for_real_ocr


ROOT = Path(__file__).resolve().parents[1]


class PublicCliTests(unittest.TestCase):
    def test_help_exposes_external_excel_only_cli(self) -> None:
        completed = _run_cli("--help")

        self.assertEqual(completed.returncode, 0, completed.stderr)
        self.assertIn("标准模板 Excel", completed.stdout)
        self.assertIn("--standard", completed.stdout)
        self.assertIn("--image", completed.stdout)
        self.assertIn("--output-dir", completed.stdout)
        self.assertIn("--publish-cos", completed.stdout)
        self.assertIn("--cos-dry-run", completed.stdout)
        self.assertNotIn("--standard-manifest", completed.stdout)
        self.assertNotIn(".pdf", completed.stdout)

    def test_interactive_ocr_token_sets_current_process_env(self) -> None:
        original_key = os.environ.pop("PPOCRV6_API_KEY", None)
        original_token = os.environ.pop("PPOCRV6_TOKEN", None)
        try:
            with patch("sys.stdin.isatty", return_value=True), patch("getpass.getpass", return_value="token-123"):
                _ensure_ocr_token_for_real_ocr(None)

            self.assertEqual(os.environ["PPOCRV6_API_KEY"], "token-123")
        finally:
            os.environ.pop("PPOCRV6_API_KEY", None)
            if original_key is not None:
                os.environ["PPOCRV6_API_KEY"] = original_key
            if original_token is not None:
                os.environ["PPOCRV6_TOKEN"] = original_token

    def test_missing_ocr_token_in_non_interactive_mode_fails_before_real_ocr(self) -> None:
        original_key = os.environ.pop("PPOCRV6_API_KEY", None)
        original_token = os.environ.pop("PPOCRV6_TOKEN", None)
        try:
            with patch("sys.stdin.isatty", return_value=False):
                with self.assertRaises(ConsistencyPipelineError) as context:
                    _ensure_ocr_token_for_real_ocr(None)

            self.assertEqual(context.exception.error_type, "MissingOcrTokenError")
            self.assertIn("缺少 OCR token", str(context.exception))
        finally:
            if original_key is not None:
                os.environ["PPOCRV6_API_KEY"] = original_key
            if original_token is not None:
                os.environ["PPOCRV6_TOKEN"] = original_token

    def test_cli_generates_report_with_offline_ocr_fixture(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            standard_path = temp_path / "standard.xlsx"
            image_path = temp_path / "package.png"
            ocr_fixture_path = temp_path / "ocr.json"
            output_dir = temp_path / "report"
            _write_standard_xlsx_fixture(standard_path)
            image_path.write_bytes(base64.b64decode(_BLANK_PNG_BASE64))
            _write_ocr_fixture(ocr_fixture_path)

            completed = _run_cli(
                "--standard",
                str(standard_path),
                "--image",
                str(image_path),
                "--ocr-fixture",
                str(ocr_fixture_path),
                "--output-dir",
                str(output_dir),
            )

            self.assertEqual(completed.returncode, 0, completed.stderr)
            for artifact_name in (
                "standard_structure/standard_items.json",
                "standard_structure/quality_report.json",
                "comparison_result.json",
                "package_ocr_lines.json",
                "package_ocr_quality_report.json",
                "result_preview.html",
                "pipeline_summary.json",
            ):
                self.assertTrue((output_dir / artifact_name).exists(), artifact_name)

            summary = json.loads((output_dir / "pipeline_summary.json").read_text(encoding="utf-8"))
            self.assertEqual(summary["status"], "completed")
            self.assertEqual(summary["stages"]["standard_structure"]["status"], "pass")
            self.assertEqual(summary["stages"]["package_image_comparison"]["status"], "pass")

    def test_cli_can_build_cos_dry_run_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            standard_path = temp_path / "standard.xlsx"
            image_path = temp_path / "package.png"
            ocr_fixture_path = temp_path / "ocr.json"
            output_dir = temp_path / "report"
            _write_standard_xlsx_fixture(standard_path)
            image_path.write_bytes(base64.b64decode(_BLANK_PNG_BASE64))
            _write_ocr_fixture(ocr_fixture_path)

            completed = _run_cli(
                "--standard",
                str(standard_path),
                "--image",
                str(image_path),
                "--ocr-fixture",
                str(ocr_fixture_path),
                "--output-dir",
                str(output_dir),
                "--publish-cos",
                "--cos-dry-run",
                "--cos-key-prefix",
                "test/{run_id}",
                env_overrides=_fake_cos_env(),
            )

            self.assertEqual(completed.returncode, 0, completed.stderr)
            self.assertTrue((output_dir / "artifacts/06_publish/cos_upload_result.json").exists())
            self.assertTrue((output_dir / "artifacts/06_publish/public_bundle/result_preview.html").exists())
            self.assertTrue(any((output_dir / "artifacts/06_publish/public_bundle").glob("package_image.*")))
            summary = json.loads((output_dir / "pipeline_summary.json").read_text(encoding="utf-8"))
            public_summary_text = (output_dir / "artifacts/06_publish/public_bundle/pipeline_summary.public.json").read_text(encoding="utf-8")
            self.assertEqual(summary["stages"]["publish"]["status"], "pass")
            self.assertEqual(summary["publish"]["status"], "success")
            self.assertTrue(summary["publish"]["dry_run"])
            self.assertIn("https://cdn.example.test/test/pkg_consistency_", summary["key_artifacts"]["published_report_html"])
            self.assertNotIn(str(temp_path), public_summary_text)
            self.assertNotIn("fake-secret", public_summary_text)


def _fake_cos_env() -> dict[str, str]:
    return {
        "PACKAGING_COS_SECRET_ID": "fake-id",
        "PACKAGING_COS_SECRET_KEY": "fake-secret",
        "PACKAGING_COS_BUCKET_URL": "https://cos.example.test?bucket=test-bucket&region=ap-guangzhou",
        "PACKAGING_COS_CDN_DOMAIN": "https://cdn.example.test",
    }


def _run_cli(*args: str, env_overrides: dict[str, str] | None = None) -> subprocess.CompletedProcess[str]:
    env = dict(os.environ)
    env.pop("PPOCRV6_API_KEY", None)
    if env_overrides:
        env.update(env_overrides)
    env["PYTHONPATH"] = os.pathsep.join(part for part in (str(ROOT / "src"), env.get("PYTHONPATH", "")) if part)
    return subprocess.run(
        [sys.executable, str(ROOT / "scripts" / "check_package_consistency.py"), *args],
        cwd=ROOT,
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        check=False,
    )


def _write_standard_xlsx_fixture(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    main = workbook.create_sheet("标签主文字")
    main.append(["字段编码", "标签项目", "客户填写原文", "备注"])
    main.append(["MAIN_PRODUCT_NAME", "产品名称", "广式粽子", ""])

    enterprise = workbook.create_sheet("企业信息")
    enterprise.append(["记录编号", "企业角色", "企业名称原文", "地址原文", "邮政编码", "备注"])
    enterprise.append(["E1", "委托方", "示例食品有限公司", "广州市示例路1号", "510000", ""])

    nutrition_list = workbook.create_sheet("营养表清单")
    nutrition_list.append(["营养表编号", "表名/适用范围", "底部备注/脚注原文", "备注"])
    nutrition_list.append(["N1", "营养成分表", "", ""])

    nutrition_items = workbook.create_sheet("营养项目明细")
    nutrition_items.append(["营养表编号", "项目", "每 100 克", "营养素参考值%", "备注"])
    nutrition_items.append(["N1", "能量", "100 千焦", "1%", ""])

    content = workbook.create_sheet("多内容物组合装")
    content.append(["内容物编号", "内容物名称", "净含量/数量", "产品分类", "配料原文", "备注"])
    content.append(["C1", "示例粽子", "100 克", "熟制粽子", "糯米", ""])

    other = workbook.create_sheet("其他标签文字")
    other.append(["记录编号", "标签文字类型", "客户填写原文", "备注"])
    other.append(["O1", "提示语", "请按包装说明食用", ""])

    workbook.save(path)


def _write_ocr_fixture(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "result": {
                    "ocrResults": [
                        {
                            "prunedResult": {
                                "input_img_shape": [1, 1, 3],
                                "rec_texts": ["广式粽子"],
                                "rec_scores": [0.99],
                                "dt_polys": [[[0, 0], [1, 0], [1, 1], [0, 1]]],
                            }
                        }
                    ]
                }
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


_BLANK_PNG_BASE64 = "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8/5+hHgAHggJ/PchI7wAAAABJRU5ErkJggg=="


if __name__ == "__main__":
    unittest.main()
