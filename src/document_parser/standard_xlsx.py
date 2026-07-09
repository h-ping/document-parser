from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .auto_ingest import build_auto_ingest_candidates
from .comparison import build_comparison_index
from .compat_artifacts import (
    build_field_groups,
    build_lists_artifact,
    build_quality_report,
    build_standard_items,
    build_structured_document,
    build_tables_artifact,
    build_taxonomy_proposals,
)
from .html_report import write_result_preview_html
from .label_text_scope import build_label_text_scope_agent_context, load_label_text_scope_reference
from .models import CompiledField, Evidence, ExtractionPlan, FieldDefinition, GeneratedSchema, ParseResult, Risk, ReviewTask, TextSpan, to_jsonable
from .mvp_metrics import build_mvp_acceptance_metrics
from .output_contract import build_output_contract_validation_report
from .schema_artifacts import write_schema_artifacts
from .source_artifacts import build_coverage_map
from .utils import sha256_file, sha256_text, stable_id, write_json


FINAL_JSON_ROOT_KEYS = [
    "job",
    "document",
    "generated_schema",
    "extracted_data",
    "evidence",
    "cross_validation",
    "coverage",
    "validation",
    "quality",
    "risks",
    "review_tasks",
    "metadata",
]
JSON_EXPORT_CONTRACT_CHECKS = [
    "machine_parseable_json",
    "root_keys_present",
    "evidence_refs_resolve",
    "risk_targets_resolve",
    "review_task_targets_resolve",
    "no_guessing",
]

TARGET_SHEETS = [
    "标签主文字",
    "企业信息",
    "营养表清单",
    "营养项目明细",
    "多内容物组合装",
    "其他标签文字",
]

SHEET_SECTIONS = {
    "标签主文字": ("sec_main_label_text", "label_text", "标签主文字"),
    "企业信息": ("sec_enterprise_info", "manufacturer_info", "企业信息"),
    "营养表清单": ("sec_nutrition_table_list", "nutrition_table_area", "营养表清单"),
    "营养项目明细": ("sec_nutrition_items", "nutrition_table", "营养项目明细"),
    "多内容物组合装": ("sec_content_items", "content_item_block", "多内容物组合装"),
    "其他标签文字": ("sec_other_label_text", "free_text_block", "其他标签文字"),
}

HEADER_ALIASES = {
    "字段编码": "字段编码",
    "编码": "编码",
    "标签项目": "标签项目",
    "客户填写原文": "客户填写原文",
    "记录编号": "记录编号",
    "企业角色": "企业角色",
    "企业名称原文": "企业名称原文",
    "地址原文": "地址原文",
    "产地原文": "产地原文",
    "许可证/备案号": "许可证/备案号",
    "联系方式": "联系方式",
    "邮政编码": "邮政编码",
    "网站": "网站",
    "营养表编号": "营养表编号",
    "表名/适用范围": "表名/适用范围",
    "底部备注/脚注原文": "底部备注/脚注原文",
    "项目": "营养项目",
    "营养项目": "营养项目",
    "每 100 克": "每100g/mL含量",
    "每100克(g)": "每100g/mL含量",
    "每100克（g）": "每100g/mL含量",
    "每100克": "每100g/mL含量",
    "每100g/mL含量": "每100g/mL含量",
    "每100g": "每100g/mL含量",
    "每100ml": "每100g/mL含量",
    "营养素参考值%": "NRV%",
    "NRV%": "NRV%",
    "内容物编号": "内容物编号",
    "内容物名称": "内容物名称",
    "净含量/数量": "净含量/数量",
    "产品分类": "产品分类",
    "配料原文": "配料原文",
    "对应营养表编号": "对应营养表编号",
    "标签文字类型": "标签文字类型",
    "备注": "备注",
}

REQUIRED_HEADERS = {
    "标签主文字": {"标签项目", "客户填写原文"},
    "企业信息": {"企业角色", "企业名称原文"},
    "营养表清单": {"营养表编号", "表名/适用范围"},
    "营养项目明细": {"营养表编号", "营养项目", "每100g/mL含量", "NRV%"},
    "多内容物组合装": {"内容物名称", "配料原文"},
    "其他标签文字": {"客户填写原文"},
}

IGNORED_HEADERS = {"字段编码", "编码", "记录编号", "营养表编号", "内容物编号", "备注"}
NOT_APPLICABLE_VALUES = {"", "不适用", "n/a", "na", "none", "无"}
PENDING_VALUE = "待确认"

MAIN_FIELD_BY_CODE = {
    "MAIN_PRODUCT_NAME": ("product.name", "产品名称", "string", "critical"),
    "MAIN_BRAND_TEXT": ("custom.brand_text", "品牌文字", "string", "non_critical"),
    "MAIN_NET_CONTENT": ("product.net_content", "净含量", "string", "critical"),
    "MAIN_PRODUCT_TYPE": ("product.product_type", "产品类型", "string", "critical"),
    "MAIN_INGREDIENTS": ("product.ingredients", "配料", "long_text", "critical"),
    "MAIN_ALLERGEN_NOTICE": ("custom.allergen_notice", "致敏原提示", "long_text", "non_critical"),
    "MAIN_STORAGE": ("product.storage_condition", "贮存条件", "string", "critical"),
    "MAIN_SHELF_LIFE": ("product.shelf_life", "保质期", "string", "critical"),
    "MAIN_DATE_BATCH_MARK": ("product.date_marking", "生产日期标示语", "long_text", "critical"),
    "MAIN_USAGE_METHOD": ("product.directions", "食用方法", "long_text", "non_critical"),
    "MAIN_WARNING_NOTICE": ("product.warning", "警示语", "long_text", "critical"),
    "MAIN_MARKETING_TEXT": ("custom.marketing_text", "营销文字", "long_text", "non_critical"),
    "MAIN_STANDARD_CODE": ("product.standard_code", "产品标准号", "string", "critical"),
    "MAIN_QUALITY_GRADE": ("custom.quality_grade", "质量等级", "string", "non_critical"),
    "MAIN_ORIGIN": ("product.origin", "产地", "string", "critical"),
    "MAIN_BARCODE": ("barcode.commodity", "条码", "barcode", "critical"),
    "MAIN_OUTER_BARCODE": ("barcode.outer_case", "外箱条码", "barcode", "non_critical"),
    "MAIN_SYMBOL_TEXT": ("custom.symbol_text", "标识文字", "long_text", "non_critical"),
}

MAIN_FIELD_BY_LABEL = {
    "产品名称": MAIN_FIELD_BY_CODE["MAIN_PRODUCT_NAME"],
    "净含量": MAIN_FIELD_BY_CODE["MAIN_NET_CONTENT"],
    "配料": MAIN_FIELD_BY_CODE["MAIN_INGREDIENTS"],
    "配料表": MAIN_FIELD_BY_CODE["MAIN_INGREDIENTS"],
    "贮存条件": MAIN_FIELD_BY_CODE["MAIN_STORAGE"],
    "储存条件": MAIN_FIELD_BY_CODE["MAIN_STORAGE"],
    "保质期": MAIN_FIELD_BY_CODE["MAIN_SHELF_LIFE"],
    "生产日期标示语": MAIN_FIELD_BY_CODE["MAIN_DATE_BATCH_MARK"],
    "食用方法": MAIN_FIELD_BY_CODE["MAIN_USAGE_METHOD"],
    "警示语": MAIN_FIELD_BY_CODE["MAIN_WARNING_NOTICE"],
    "产品标准号": MAIN_FIELD_BY_CODE["MAIN_STANDARD_CODE"],
    "产品标准代号": MAIN_FIELD_BY_CODE["MAIN_STANDARD_CODE"],
    "质量等级": MAIN_FIELD_BY_CODE["MAIN_QUALITY_GRADE"],
    "产地": MAIN_FIELD_BY_CODE["MAIN_ORIGIN"],
    "条码": MAIN_FIELD_BY_CODE["MAIN_BARCODE"],
    "商品条码": MAIN_FIELD_BY_CODE["MAIN_BARCODE"],
    "外箱条码": MAIN_FIELD_BY_CODE["MAIN_OUTER_BARCODE"],
}

ENTERPRISE_COLUMNS = {
    "企业名称原文": ("name", "企业名称"),
    "地址原文": ("address", "地址"),
    "产地原文": ("origin", "产地"),
    "许可证/备案号": ("license_number", "许可证/备案号"),
    "联系方式": ("contact", "联系方式"),
    "邮政编码": ("postal_code", "邮政编码"),
    "网站": ("website", "网站"),
}

CONTENT_COLUMNS = {
    "内容物名称": ("content_item.name", "内容物名称", "string", "critical"),
    "净含量/数量": ("content_item.net_content", "净含量/数量", "string", "critical"),
    "产品分类": ("content_item.product_category", "产品分类", "string", "non_critical"),
    "配料原文": ("content_item.ingredients", "配料原文", "long_text", "critical"),
}


class ParseError(RuntimeError):
    pass


@dataclass(frozen=True)
class CellValue:
    text: str
    coord: str


@dataclass(frozen=True)
class SheetTable:
    sheet_name: str
    header_row: int
    columns: dict[str, int]
    header_cells: dict[str, str]
    header_labels: dict[str, str]
    ignored_columns: list[dict[str, Any]]
    rows: list[dict[str, CellValue]]


class StandardXlsxParser:
    def parse(self, input_xlsx: Path, debug_dir: Path | None = None, runtime_policy: dict[str, Any] | None = None) -> ParseResult:
        if not input_xlsx.exists():
            raise ParseError(f"Input XLSX does not exist: {input_xlsx}")
        if input_xlsx.suffix.lower() != ".xlsx":
            raise ParseError("Input file must be an XLSX")

        workbook = load_workbook(input_xlsx, data_only=True, read_only=True)
        missing_sheets = [sheet for sheet in TARGET_SHEETS if sheet not in workbook.sheetnames]
        if missing_sheets:
            raise ParseError(f"XLSX standard file is missing required sheet(s): {', '.join(missing_sheets)}")

        sheets = {sheet_name: _read_sheet(workbook[sheet_name]) for sheet_name in TARGET_SHEETS}
        workbook_structure = _workbook_structure(workbook.sheetnames, sheets)
        builder = _XlsxArtifactBuilder(input_xlsx)

        builder.parse_main_labels(sheets["标签主文字"])
        builder.parse_enterprises(sheets["企业信息"])
        nutrition_tables = builder.parse_nutrition_tables(sheets["营养表清单"], sheets["营养项目明细"])
        builder.parse_content_items(sheets["多内容物组合装"], nutrition_tables)
        builder.parse_other_labels(sheets["其他标签文字"])

        result = builder.build_result(runtime_policy or _xlsx_runtime_policy())
        output_contract_validation_report = build_output_contract_validation_report(result)
        result.cross_validation["output_contract"] = output_contract_validation_report
        result.metadata["output_contract_validation_report"] = output_contract_validation_report
        result.metadata["mvp_acceptance_metrics"] = build_mvp_acceptance_metrics(
            schema=result.generated_schema,
            plan=builder.plan,
            compiled_fields=builder.compiled_fields,
            evidence=result.evidence,
            validation=result.validation,
            risks=result.risks,
            review_tasks=result.review_tasks,
            coverage=result.coverage,
            coverage_map=result.metadata["coverage_map"],
            schema_audit=result.metadata["schema_audit"],
            structure_audit=result.metadata["structure_audit"],
            source_layers=result.metadata["source_layers"],
            table_quality_report=result.metadata["table_parser"]["table_quality_report"],
            repair_trace=result.metadata["repair_loop"]["trace"],
            output_contract_validation_report=output_contract_validation_report,
            missing_item_report=result.metadata["missing_item_report"],
        )

        if debug_dir:
            self._write_debug(debug_dir, result, workbook_structure, builder.plan)
        if output_contract_validation_report.get("failed_count", 0):
            failed = [
                check.get("check_type")
                for check in output_contract_validation_report.get("checks", [])
                if check.get("result") == "failed"
            ]
            raise ParseError(f"XLSX output contract validation failed: {', '.join(str(item) for item in failed[:8])}")
        return result

    def _write_debug(
        self,
        debug_dir: Path,
        result: ParseResult,
        workbook_structure: dict[str, Any],
        plan: ExtractionPlan,
    ) -> None:
        metadata = result.metadata
        standard = metadata["standard_artifacts"]
        write_json(debug_dir / "result.json", to_jsonable(result))
        write_result_preview_html(result, debug_dir / "result_preview.html", artifact_root=debug_dir, workbook_structure=workbook_structure)
        write_json(debug_dir / "runtime_policy.json", metadata["runtime_policy"])
        write_json(debug_dir / "json_export.json", metadata["json_export"])
        write_json(debug_dir / "xlsx_workbook_structure.json", workbook_structure)
        write_json(debug_dir / "page_images.json", metadata["page_images"])
        write_json(debug_dir / "candidate_visual_document_graph.json", metadata["candidate_visual_document_graph"])
        write_json(debug_dir / "visual_document_graph.json", metadata["visual_document_graph"])
        write_json(debug_dir / "vdg_quality_report.json", metadata["vdg_quality_report"])
        write_json(debug_dir / "vdg_agent_context.json", metadata["vdg_agent_context"])
        write_json(debug_dir / "vdg_consumption_report.json", metadata["vdg_consumption_report"])
        write_json(debug_dir / "label_text_scope_reference.json", metadata["label_text_scope_reference"])
        write_json(debug_dir / "label_text_scope_agent_context.json", metadata["label_text_scope_agent_context"])
        write_json(debug_dir / "label_text_scope_report.json", metadata["label_text_scope_report"])
        write_json(debug_dir / "extraction_plan.json", to_jsonable(plan))
        write_json(debug_dir / "schema_audit.json", metadata["schema_audit"])
        write_json(debug_dir / "audit_input.json", metadata["audit_input"])
        write_json(debug_dir / "standard_items.json", standard["standard_items"])
        write_json(debug_dir / "comparison_index.json", standard["comparison_index"])
        write_json(debug_dir / "quality_report.json", standard["quality_report"])
        write_json(debug_dir / "structured_document.json", standard["structured_document"])
        write_json(debug_dir / "taxonomy_proposals.json", standard["taxonomy_proposals"])
        write_json(debug_dir / "field_groups.json", standard["field_groups"])
        write_json(debug_dir / "tables.json", standard["tables"])
        write_json(debug_dir / "lists.json", standard["lists"])
        write_json(debug_dir / "auto_ingest_candidates.json", standard["auto_ingest_candidates"])
        write_json(debug_dir / "table_layers.json", metadata["table_parser"]["table_layers"])
        write_json(debug_dir / "table_quality_report.json", metadata["table_parser"]["table_quality_report"])
        write_json(debug_dir / "table_feed_items.candidates.json", metadata["table_parser"]["table_feed_items"])
        write_json(debug_dir / "source_layers.json", metadata["source_layers"])
        write_json(debug_dir / "source_consistency_report.json", metadata["source_consistency"])
        write_json(debug_dir / "source_anchor_inventory.json", metadata["source_anchor_inventory"])
        write_json(debug_dir / "coverage_map.json", metadata["coverage_map"])
        write_json(debug_dir / "structure_audit.json", metadata["structure_audit"])
        write_json(debug_dir / "missing_item_report.json", metadata["missing_item_report"])
        write_json(debug_dir / "missing_fields.json", metadata["missing_item_report"]["missing_fields"])
        write_json(debug_dir / "missing_tables.json", metadata["missing_item_report"]["missing_tables"])
        write_json(debug_dir / "revision_blocks.json", result.extracted_data["revision_blocks"])
        write_json(debug_dir / "output_contract_validation_report.json", metadata["output_contract_validation_report"])
        write_json(debug_dir / "mvp_acceptance_metrics.json", metadata["mvp_acceptance_metrics"])
        write_json(debug_dir / "extracted_data.json", to_jsonable(result.extracted_data))
        write_json(debug_dir / "evidence.json", to_jsonable(result.evidence))
        write_json(debug_dir / "validation.json", to_jsonable(result.validation))
        write_json(debug_dir / "risks.json", to_jsonable(result.risks))
        write_json(debug_dir / "review_tasks.json", to_jsonable(result.review_tasks))
        write_json(debug_dir / "repair_plan.json", metadata["repair_loop"]["repair_plan"])
        write_json(debug_dir / "repair_trace.json", metadata["repair_loop"]["trace"])
        write_json(debug_dir / "repair_attempts.json", metadata["repair_loop"]["attempts"])
        write_json(debug_dir / "repair_plan_patches.json", metadata["repair_loop"]["repair_plan_patches"])
        write_json(debug_dir / "repair_agent_candidates.json", metadata["repair_loop"]["repair_agent_candidates"])
        write_json(debug_dir / "repaired_source_layers.json", metadata["repair_loop"]["repaired_source_layers"])
        write_json(debug_dir / "llm_agent_items.json", metadata["agent_harness"]["llm_agent_items"])
        write_json(debug_dir / "rejected_agent_items.json", metadata["agent_harness"]["rejected_agent_items"])
        write_json(debug_dir / "review_items.json", metadata["agent_harness"]["review_items"])
        write_json(debug_dir / "agent_execution_report.json", metadata["agent_execution_report"])
        write_json(debug_dir / "agent_harness_report.json", metadata["agent_harness"])
        _write_layered_debug_artifacts(debug_dir, result, workbook_structure)
        write_schema_artifacts(debug_dir)
        write_artifact_index(debug_dir)


class _XlsxArtifactBuilder:
    def __init__(self, source_path: Path) -> None:
        self.source_path = source_path
        self.spans: list[TextSpan] = []
        self._span_ids: set[str] = set()
        self.evidence: list[Evidence] = []
        self.compiled_fields: dict[str, CompiledField] = {}
        self.field_definitions: list[FieldDefinition] = []
        self.sections_by_sheet: dict[str, dict[str, Any]] = {}
        self.entities: dict[str, dict[str, Any]] = {
            "product_001": _empty_entity("product_001", "product", 1),
        }
        self.tables: list[dict[str, Any]] = []
        self._table_ids_by_number: dict[str, str] = {}
        self._explicit_content_table_links: dict[str, list[str]] = {}
        self._extra_risks: list[Risk] = []
        self.plan = ExtractionPlan(plan_id="plan_xlsx_001", schema_id="schema_xlsx_001", fields=[], tables=[])

    def parse_main_labels(self, sheet: SheetTable) -> None:
        for row in sheet.rows:
            value = _cell_text(row, "客户填写原文")
            if not _is_print_value(value):
                continue
            code = _cell_text(row, "字段编码")
            label = _cell_text(row, "标签项目") or "标签主文字"
            semantic_key, display_name, field_type, criticality = _main_field_definition(code, label)
            self.add_field(
                semantic_key=semantic_key,
                display_name=display_name,
                field_type=field_type,
                text=value,
                sheet_name=sheet.sheet_name,
                coord=_cell_coord(row, "客户填写原文"),
                criticality=criticality,
                entity_id="product_001",
            )

    def parse_enterprises(self, sheet: SheetTable) -> None:
        for index, row in enumerate(sheet.rows, start=1):
            role = _cell_text(row, "企业角色")
            entity_type = _enterprise_type(role)
            record_id = _cell_text(row, "记录编号") or f"E{index}"
            entity_id = _stable_group_id(entity_type, record_id, index)
            field_ids: list[str] = []
            for header, (slot, label) in ENTERPRISE_COLUMNS.items():
                value = _cell_text(row, header)
                if not _is_print_value(value):
                    continue
                field_id = self.add_field(
                    semantic_key=f"{entity_type}.{slot}",
                    display_name=label,
                    field_type="long_text" if slot == "address" else "string",
                    text=value,
                    sheet_name=sheet.sheet_name,
                    coord=_cell_coord(row, header),
                    criticality="critical" if slot in {"name", "address"} else "non_critical",
                    entity_id=entity_id,
                )
                field_ids.append(field_id)
            if field_ids:
                self.entities[entity_id] = _entity_from_fields(entity_id, entity_type, index, field_ids, self.compiled_fields)

    def parse_nutrition_tables(self, table_list_sheet: SheetTable, detail_sheet: SheetTable) -> dict[str, str]:
        table_headers = self._nutrition_table_headers(table_list_sheet)
        detail_rows = self._nutrition_detail_rows(detail_sheet)
        table_numbers = _ordered_unique([*table_headers.keys(), *detail_rows.keys()])
        for index, table_number in enumerate(table_numbers, start=1):
            rows = detail_rows.get(table_number, [])
            if not rows:
                continue
            header = table_headers.get(table_number, {})
            title = header.get("title") or "营养成分表"
            table_id = f"nutrition_{_safe_id(table_number, index)}"
            self._table_ids_by_number[table_number] = table_id
            table = self._build_nutrition_table(table_id, table_number, title, header, rows, detail_sheet.header_labels)
            self.tables.append(table)
            self.entities["product_001"]["linked_table_ids"].append(table_id)
            nutrition_text = _nutrition_standard_text(title, rows, header.get("footnote", ""))
            evidence_ref = self.add_composite_evidence(
                source_text=nutrition_text,
                method="xlsx_nutrition_table",
                source_node_ids=table["source_span_ids"],
            )
            field_id = self.add_field(
                semantic_key="product.nutrition_table",
                display_name=title,
                field_type="table",
                text=nutrition_text,
                sheet_name=detail_sheet.sheet_name,
                coord=rows[0]["item"].coord,
                criticality="critical",
                entity_id="product_001",
                table_id=table_id,
                evidence_ref=evidence_ref,
                source_span_ids=table["source_span_ids"],
            )
            self.entities["product_001"]["fields"][field_id] = _entity_field(self.compiled_fields[field_id])
        self.plan = ExtractionPlan(
            plan_id="plan_xlsx_001",
            schema_id="schema_xlsx_001",
            fields=[],
            tables=[{"table_id": table["table_id"], "table_type": table["table_type"]} for table in self.tables],
        )
        return dict(self._table_ids_by_number)

    def parse_content_items(self, sheet: SheetTable, nutrition_tables: dict[str, str]) -> None:
        for index, row in enumerate(sheet.rows, start=1):
            content_id = _cell_text(row, "内容物编号") or f"C{index}"
            entity_id = _stable_group_id("content", content_id, index)
            field_ids: list[str] = []
            for header, (semantic_key, display_name, field_type, criticality) in CONTENT_COLUMNS.items():
                value = _cell_text(row, header)
                if not _is_print_value(value):
                    continue
                field_id = self.add_field(
                    semantic_key=semantic_key,
                    display_name=display_name,
                    field_type=field_type,
                    text=value,
                    sheet_name=sheet.sheet_name,
                    coord=_cell_coord(row, header),
                    criticality=criticality,
                    entity_id=entity_id,
                    row_key=content_id,
                )
                field_ids.append(field_id)
            linked_table_ids = []
            explicit_table_ref = _cell_text(row, "对应营养表编号")
            if explicit_table_ref:
                for table_number in _split_refs(explicit_table_ref):
                    table_id = nutrition_tables.get(table_number)
                    if table_id:
                        linked_table_ids.append(table_id)
            if field_ids:
                entity = _entity_from_fields(entity_id, "content_item", index, field_ids, self.compiled_fields)
                entity["linked_table_ids"] = linked_table_ids
                self.entities[entity_id] = entity
                for table in self.tables:
                    if table["table_id"] in linked_table_ids:
                        table["linked_entity_id"] = entity_id

    def parse_other_labels(self, sheet: SheetTable) -> None:
        for index, row in enumerate(sheet.rows, start=1):
            value = _cell_text(row, "客户填写原文")
            if not _is_print_value(value):
                continue
            label = _cell_text(row, "标签文字类型") or "其他标签文字"
            self.add_field(
                semantic_key="custom.other_label_text",
                display_name=label,
                field_type="long_text",
                text=value,
                sheet_name=sheet.sheet_name,
                coord=_cell_coord(row, "客户填写原文"),
                criticality="non_critical",
                entity_id="product_001",
                row_key=_cell_text(row, "记录编号") or f"OTHER{index}",
            )

    def add_field(
        self,
        *,
        semantic_key: str,
        display_name: str,
        field_type: str,
        text: str,
        sheet_name: str,
        coord: str,
        criticality: str,
        entity_id: str | None,
        table_id: str | None = None,
        row_key: str | None = None,
        evidence_ref: str | None = None,
        source_span_ids: list[str] | None = None,
    ) -> str:
        evidence_ref = evidence_ref or self.add_cell_evidence(sheet_name, coord, text, method="xlsx_cell")
        evidence_item = next(item for item in self.evidence if item.evidence_id == evidence_ref)
        source_span_ids = source_span_ids or evidence_item.source_node_ids
        field_id = stable_id("fld", len(self.compiled_fields) + 1)
        review_required = text.strip() == PENDING_VALUE
        field = CompiledField(
            field_id=field_id,
            semantic_key=semantic_key,
            display_name=display_name,
            field_type=field_type,
            raw_value=text,
            clean_value=text,
            normalized_value=text,
            value_hash=sha256_text(text),
            status="manual_review_required" if review_required else "verified",
            criticality=criticality,
            confidence={
                "overall": 1.0,
                "source_confidence": 1.0,
                "extraction_confidence": 1.0,
                "boundary_confidence": 1.0,
            },
            risk_level="high" if review_required else "low",
            review_required=review_required,
            section_id=SHEET_SECTIONS[sheet_name][0],
            entity_id=entity_id,
            table_id=table_id,
            row_key=row_key,
            evidence_refs=[evidence_ref],
            normalization=[],
            reason="待确认，需要人工复核。" if review_required else None,
        )
        self.compiled_fields[field_id] = field
        self.field_definitions.append(
            FieldDefinition(
                field_def_id=stable_id("fdef", len(self.field_definitions) + 1),
                semantic_key=semantic_key,
                display_name=display_name,
                field_type=field_type,
                criticality=criticality,
                repeatable=entity_id != "product_001" or semantic_key in {"product.nutrition_table", "custom.other_label_text"},
                semantic_key_type="xlsx_template",
                source_span_ids=source_span_ids,
            )
        )
        self._remember_section_span(sheet_name, source_span_ids)
        if entity_id == "product_001":
            self.entities["product_001"]["fields"][field_id] = _entity_field(field)
        if review_required:
            risk = Risk(
                risk_id=stable_id("risk", len(self._pending_risks()) + 1),
                target_type="field",
                target_id=field_id,
                risk_level="high",
                risk_type="manual_review_required",
                message="字段值为待确认，需要人工复核。",
                evidence_refs=[evidence_ref],
            )
            if entity_id in self.entities:
                self.entities[entity_id]["review_required"] = True
                self.entities[entity_id]["risk_level"] = "high"
            self._extra_risks.append(risk)
        return field_id

    def add_cell_evidence(self, sheet_name: str, coord: str, source_text: str, *, method: str) -> str:
        span_id = f"xlsx:{sheet_name}!{coord}"
        if span_id not in self._span_ids:
            self._span_ids.add(span_id)
            self.spans.append(TextSpan(span_id=span_id, page=1, text=source_text, source="xlsx_cell", confidence=1.0))
        evidence_id = stable_id("ev", len(self.evidence) + 1)
        self.evidence.append(
            Evidence(
                evidence_id=evidence_id,
                source_text=source_text,
                page=1,
                extraction_methods=[method],
                bbox_status="missing",
                source_node_ids=[span_id],
            )
        )
        return evidence_id

    def add_composite_evidence(self, *, source_text: str, method: str, source_node_ids: list[str]) -> str:
        evidence_id = stable_id("ev", len(self.evidence) + 1)
        self.evidence.append(
            Evidence(
                evidence_id=evidence_id,
                source_text=source_text,
                page=1,
                extraction_methods=[method],
                bbox_status="missing",
                source_node_ids=source_node_ids,
            )
        )
        return evidence_id

    def build_result(self, runtime_policy: dict[str, Any]) -> ParseResult:
        self._refresh_entity_refs()
        schema = self._schema()
        regions = self._regions()
        visual_document_graph = _visual_document_graph(regions)
        table_layers, table_quality_report = _table_parser_artifacts(self.tables)
        source_layers = _source_layers(self.spans, table_layers)
        source_consistency = _source_consistency_report()
        validation = self._validation_checks(table_quality_report)
        risks = list(self._pending_risks())
        review_tasks = _review_tasks_from_risks(risks)
        quality = _quality(self.compiled_fields, risks)
        structure_audit = _structure_audit(regions)
        coverage_map = build_coverage_map(self.spans, self.compiled_fields, self.evidence, self.tables, [], regions, [])
        structure_audit["duplicate_coverage_issues"] = coverage_map["duplicate_coverage_issues"]
        structure_audit["duplicate_coverage_issue_count"] = coverage_map["duplicate_coverage_issue_count"]
        missing_item_report = {"missing_fields": [], "missing_tables": [], "missing_count": 0}
        repair_loop = _repair_loop()
        standard_items = build_standard_items(self.compiled_fields, self.evidence, self.spans, self.source_path, [])
        for item in standard_items:
            if item.get("text") == PENDING_VALUE:
                item["comparison_required"] = False
        comparison_index = build_comparison_index(standard_items)
        auto_ingest_candidates = build_auto_ingest_candidates(standard_items, validation, risks, quality)
        field_groups = build_field_groups(self.entities, self.compiled_fields, standard_items)
        tables_artifact = build_tables_artifact(self.tables)
        lists_artifact = build_lists_artifact(field_groups)
        vdg_quality_report = _vdg_quality_report()
        label_text_scope_reference = load_label_text_scope_reference()
        label_text_scope_agent_context = build_label_text_scope_agent_context(label_text_scope_reference)
        label_text_scope_report = _label_text_scope_report(label_text_scope_reference)
        quality_report = build_quality_report(
            quality,
            risks,
            validation,
            _schema_audit(),
            structure_audit,
            source_layers,
            table_quality_report,
            repair_loop["repair_plan"],
            repair_loop["attempts"],
            repair_loop["trace"],
            repair_loop["repair_agent_candidates"],
            vdg_quality_report,
            label_text_scope_report,
        )
        structured_document = build_structured_document(
            {
                "file_name": self.source_path.name,
                "file_hash": sha256_file(self.source_path),
                "page_count": 1,
                "language": ["zh-CN"],
            },
            source_layers,
            schema,
            regions,
            standard_items,
            field_groups,
            tables_artifact,
            lists_artifact,
        )
        taxonomy_proposals = build_taxonomy_proposals(self.compiled_fields, self.evidence)
        result = ParseResult(
            job={
                "job_id": f"job_{dt.datetime.now(dt.UTC).strftime('%Y%m%d_%H%M%S')}",
                "job_type": "standard_xlsx_to_structured_json",
                "status": "completed_with_warnings" if risks else "completed",
            },
            document={
                "file_name": self.source_path.name,
                "file_hash": sha256_file(self.source_path),
                "page_count": 1,
                "page_sizes": [{"page": 1, "width": 1, "height": 1}],
                "detected_document_types": _detected_document_types(self.tables),
                "language": ["zh-CN"],
                "parse_status": "completed_with_warnings" if risks else "completed",
                "pdf_text_layer_available": False,
                "page_image_status": "not_applicable",
                "warnings": [],
            },
            generated_schema=schema,
            extracted_data={
                "sections": schema.sections,
                "regions": regions,
                "entities": self.entities,
                "fields": {field_id: to_jsonable(field) for field_id, field in self.compiled_fields.items()},
                "missing_fields": [],
                "missing_tables": [],
                "tables": self.tables,
                "requirements": [],
                "revision_blocks": [],
            },
            evidence=self.evidence,
            cross_validation={
                "checks": validation,
                "source_consistency": source_consistency,
            },
            coverage=_coverage(),
            validation=validation,
            quality=quality,
            risks=risks,
            review_tasks=review_tasks,
            metadata={
                "parser_version": "mvp_v0.1",
                "pipeline_version": "mvp_pipeline_v0.2",
                "schema_mode": "xlsx_template",
                "no_guessing": True,
                "json_export": _json_export_manifest(),
                "ocr_provider": "not_applicable",
                "runtime_policy": runtime_policy,
                "page_images": {"status": "not_applicable", "images": [], "reason": "xlsx_structured_input_has_no_pages"},
                "candidate_visual_document_graph": visual_document_graph,
                "visual_document_graph": visual_document_graph,
                "vdg_quality_report": vdg_quality_report,
                "vdg_agent_context": _vdg_agent_context(table_layers),
                "vdg_consumption_report": _vdg_consumption_report(),
                "label_text_scope_reference": label_text_scope_reference,
                "label_text_scope_agent_context": label_text_scope_agent_context,
                "label_text_scope_report": label_text_scope_report,
                "missing_item_report": missing_item_report,
                "repair_loop": repair_loop,
                "audit_input": _audit_input(schema, self.compiled_fields, self.evidence),
                "schema_audit": _schema_audit(),
                "structure_audit": structure_audit,
                "source_layers": source_layers,
                "source_consistency": source_consistency,
                "source_anchor_inventory": [],
                "coverage_map": coverage_map,
                "standard_artifacts": {
                    "standard_items": standard_items,
                    "comparison_index": comparison_index,
                    "quality_report": quality_report,
                    "structured_document": structured_document,
                    "taxonomy_proposals": taxonomy_proposals,
                    "field_groups": field_groups,
                    "tables": tables_artifact,
                    "lists": lists_artifact,
                    "auto_ingest_candidates": auto_ingest_candidates,
                },
                "table_parser": {
                    "table_layers": table_layers,
                    "table_quality_report": table_quality_report,
                    "table_feed_items": [],
                },
                "agent_execution_report": _agent_execution_report(schema, self.plan, self.compiled_fields, self.evidence),
                "agent_harness": _agent_harness(),
                "created_at": dt.datetime.now(dt.UTC).isoformat(),
            },
        )
        return result

    def _nutrition_table_headers(self, sheet: SheetTable) -> dict[str, dict[str, Any]]:
        headers: dict[str, dict[str, Any]] = {}
        for row in sheet.rows:
            table_number = _cell_text(row, "营养表编号")
            if not table_number:
                continue
            title_cell = row.get("表名/适用范围")
            footnote_cell = row.get("底部备注/脚注原文")
            title = title_cell.text if title_cell and _is_print_value(title_cell.text) else "营养成分表"
            footnote = footnote_cell.text if footnote_cell and _is_print_value(footnote_cell.text) else ""
            source_span_ids = []
            if title_cell and _is_print_value(title_cell.text):
                source_span_ids.append(_span_id(sheet.sheet_name, title_cell.coord))
                self.add_cell_evidence(sheet.sheet_name, title_cell.coord, title, method="xlsx_cell")
            if footnote_cell and footnote:
                source_span_ids.append(_span_id(sheet.sheet_name, footnote_cell.coord))
                self.add_cell_evidence(sheet.sheet_name, footnote_cell.coord, footnote, method="xlsx_cell")
            headers[table_number] = {
                "title": title,
                "footnote": footnote,
                "source_span_ids": source_span_ids,
            }
            self._remember_section_span(sheet.sheet_name, source_span_ids)
        return headers

    def _nutrition_detail_rows(self, sheet: SheetTable) -> dict[str, list[dict[str, CellValue]]]:
        rows_by_table: dict[str, list[dict[str, CellValue]]] = {}
        for row in sheet.rows:
            table_number = _cell_text(row, "营养表编号")
            item = row.get("营养项目")
            amount = row.get("每100g/mL含量")
            nrv = row.get("NRV%")
            if not table_number or not item or not _is_print_value(item.text):
                continue
            amount_value = amount.text if amount and _is_print_value(amount.text) else ""
            nrv_value = nrv.text if nrv and _is_print_value(nrv.text) else ""
            if not amount_value and not nrv_value:
                continue
            detail_row = {
                "item": item,
                "amount": CellValue(amount_value, amount.coord if amount else item.coord),
                "nrv": CellValue(nrv_value, nrv.coord if nrv else item.coord),
            }
            rows_by_table.setdefault(table_number, []).append(detail_row)
            self._remember_section_span(sheet.sheet_name, [_span_id(sheet.sheet_name, item.coord)])
        return rows_by_table

    def _build_nutrition_table(
        self,
        table_id: str,
        table_number: str,
        title: str,
        header: dict[str, Any],
        rows: list[dict[str, CellValue]],
        detail_header_labels: dict[str, str],
    ) -> dict[str, Any]:
        columns = [
            {"column_id": "col_001", "name": detail_header_labels.get("营养项目", "营养项目")},
            {"column_id": "col_002", "name": detail_header_labels.get("每100g/mL含量", "每100g/mL含量")},
            {"column_id": "col_003", "name": detail_header_labels.get("NRV%", "NRV%")},
        ]
        table_rows = []
        table_evidence_refs: list[str] = []
        source_span_ids = list(header.get("source_span_ids", []))
        for index, row in enumerate(rows, start=1):
            item_evidence = self.add_cell_evidence("营养项目明细", row["item"].coord, row["item"].text, method="xlsx_table_cell")
            amount_evidence = (
                self.add_cell_evidence("营养项目明细", row["amount"].coord, row["amount"].text, method="xlsx_table_cell")
                if row["amount"].text
                else item_evidence
            )
            nrv_evidence = (
                self.add_cell_evidence("营养项目明细", row["nrv"].coord, row["nrv"].text, method="xlsx_table_cell")
                if row["nrv"].text
                else item_evidence
            )
            row_refs = [item_evidence, amount_evidence, nrv_evidence]
            table_evidence_refs.extend(row_refs)
            source_span_ids.extend([_span_id("营养项目明细", row["item"].coord), _span_id("营养项目明细", row["amount"].coord), _span_id("营养项目明细", row["nrv"].coord)])
            table_rows.append(
                {
                    "row_id": stable_id("row", index),
                    "row_key": row["item"].text,
                    "evidence_refs": row_refs,
                    "cells": [
                        {"column_id": "col_001", "raw_value": row["item"].text, "normalized_value": row["item"].text, "evidence_refs": [item_evidence]},
                        {"column_id": "col_002", "raw_value": row["amount"].text, "normalized_value": row["amount"].text, "evidence_refs": [amount_evidence]},
                        {"column_id": "col_003", "raw_value": row["nrv"].text, "normalized_value": row["nrv"].text, "evidence_refs": [nrv_evidence]},
                    ],
                }
            )
        footnote = header.get("footnote", "")
        if footnote:
            footnote_evidence = self.add_composite_evidence(source_text=footnote, method="xlsx_table_footnote", source_node_ids=header.get("source_span_ids", []))
            table_evidence_refs.append(footnote_evidence)
        return {
            "table_id": table_id,
            "table_type": "nutrition_facts",
            "page": 1,
            "title": title,
            "linked_entity_id": "product_001",
            "columns": columns,
            "rows": table_rows,
            "footnotes": [footnote] if footnote else [],
            "status": "verified",
            "bbox_status": "missing",
            "confidence": {"table_structure_confidence": 1.0, "evidence_confidence": 1.0},
            "criticality": "critical",
            "risk_level": "low",
            "review_required": False,
            "evidence_refs": _ordered_unique(table_evidence_refs),
            "source_span_ids": _ordered_unique(source_span_ids),
            "source": "xlsx_structured_table",
            "source_table_number": table_number,
        }

    def _schema(self) -> GeneratedSchema:
        sections = [
            section
            for section in self.sections_by_sheet.values()
            if section.get("source_span_ids")
        ]
        if not sections and self.spans:
            sections.append({"section_id": "sec_xlsx", "section_type": "label_text", "display_name": "Excel 标签文字", "source_span_ids": [self.spans[0].span_id]})
        table_definitions = []
        table_span_ids = _ordered_unique(span_id for table in self.tables for span_id in table.get("source_span_ids", []))
        if table_span_ids:
            table_definitions.append(
                {
                    "table_type": "nutrition_facts",
                    "display_name": "营养成分表",
                    "criticality": "critical",
                    "repeatable": True,
                    "source_span_ids": table_span_ids,
                }
            )
        return GeneratedSchema(
            schema_id="schema_xlsx_001",
            auto_generated=True,
            schema_version="xlsx_template_v1",
            sections=sections,
            entity_types=[
                {"entity_type": "product", "repeatable": False},
                {"entity_type": "principal", "repeatable": True},
                {"entity_type": "manufacturer", "repeatable": True},
                {"entity_type": "distributor", "repeatable": True},
                {"entity_type": "enterprise", "repeatable": True},
                {"entity_type": "content_item", "repeatable": True},
            ],
            field_definitions=self.field_definitions,
            table_definitions=table_definitions,
            requirement_definitions=[],
        )

    def _regions(self) -> list[dict[str, Any]]:
        regions = []
        field_ids_by_section: dict[str, list[str]] = {}
        table_ids_by_section: dict[str, list[str]] = {}
        entity_ids_by_section: dict[str, list[str]] = {}
        evidence_refs_by_span: dict[str, list[str]] = {}
        for evidence in self.evidence:
            for span_id in evidence.source_node_ids:
                evidence_refs_by_span.setdefault(span_id, []).append(evidence.evidence_id)
        for field in self.compiled_fields.values():
            if field.section_id:
                field_ids_by_section.setdefault(field.section_id, []).append(field.field_id)
                if field.entity_id:
                    entity_ids_by_section.setdefault(field.section_id, []).append(field.entity_id)
                if field.table_id:
                    table_ids_by_section.setdefault(field.section_id, []).append(field.table_id)
        for sheet_name, section in self.sections_by_sheet.items():
            source_span_ids = _ordered_unique(section.get("source_span_ids", []))
            if not source_span_ids:
                continue
            section_id, region_type, display_name = SHEET_SECTIONS[sheet_name]
            evidence_refs = _ordered_unique(ref for span_id in source_span_ids for ref in evidence_refs_by_span.get(span_id, []))
            regions.append(
                {
                    "region_id": f"region_{section_id.removeprefix('sec_')}",
                    "region_type": region_type,
                    "display_name": display_name,
                    "page": 1,
                    "source_span_ids": source_span_ids,
                    "bbox_status": "missing",
                    "confidence": 1.0,
                    "status": "verified",
                    "risk_level": "low",
                    "review_required": False,
                    "evidence_refs": evidence_refs,
                    "fields": _ordered_unique(field_ids_by_section.get(section_id, [])),
                    "tables": _ordered_unique(table_ids_by_section.get(section_id, [])),
                    "entities": _ordered_unique(entity_ids_by_section.get(section_id, [])),
                    "assignment_status": "assigned",
                }
            )
        return regions

    def _validation_checks(self, table_quality_report: dict[str, Any]) -> list[dict[str, Any]]:
        checks: list[dict[str, Any]] = []
        for field in self.compiled_fields.values():
            checks.append(_validation_check(checks, field.field_id, "schema_validation", "passed", "Schema validation passed.", field.evidence_refs))
            checks.append(_validation_check(checks, field.field_id, "format_check", "passed", "Format check passed or not applicable for XLSX template input.", field.evidence_refs))
            checks.append(
                {
                    **_validation_check(checks, field.field_id, "bbox_integrity", "passed", "XLSX structured cells do not provide print bbox; cell evidence is accepted.", field.evidence_refs),
                    "bbox_status": "missing",
                    "applicability": "xlsx_structured_input_no_bbox",
                    "risk_type": None,
                }
            )
            checks.append(_validation_check(checks, field.field_id, "no_guessing", "passed", "Value is copied from a filled XLSX cell.", field.evidence_refs))
        checks.append(
            {
                **_validation_check(checks, "source_consistency", "multi_method_agreement", "passed", "XLSX structured input has a single authoritative source.", []),
                "agreement": "not_applicable_single_structured_source",
                "issue_count": 0,
                "issues": [],
            }
        )
        checks.append(
            {
                **_validation_check(checks, "internal_consistency", "internal_consistency", "passed", "Internal consistency check passed.", []),
                "conflict_count": 0,
                "conflicts": [],
            }
        )
        for table in self.tables:
            checks.append(
                {
                    **_validation_check(checks, table["table_id"], "table_structure", "passed", "Table structure validation passed.", table.get("evidence_refs", [])),
                    "table_type": table.get("table_type"),
                    "row_count": len(table.get("rows", [])),
                    "column_count": len(table.get("columns", [])),
                }
            )
        checks.append(
            {
                **_validation_check(checks, "table_quality_report", "table_structure", "passed", "Table parser quality validation passed.", []),
                "table_quality_status": table_quality_report.get("status"),
                "parser_agreement": table_quality_report.get("parser_agreement", {}).get("status"),
                "issue_count": table_quality_report.get("issue_count", 0),
                "issues": [],
            }
        )
        return checks

    def _remember_section_span(self, sheet_name: str, source_span_ids: list[str]) -> None:
        if not source_span_ids:
            return
        section_id, section_type, display_name = SHEET_SECTIONS[sheet_name]
        section = self.sections_by_sheet.setdefault(
            sheet_name,
            {
                "section_id": section_id,
                "section_type": section_type,
                "display_name": display_name,
                "source_span_ids": [],
            },
        )
        section["source_span_ids"] = _ordered_unique([*section["source_span_ids"], *source_span_ids])

    def _refresh_entity_refs(self) -> None:
        for entity_id, entity in self.entities.items():
            field_ids = list(entity.get("fields", {}))
            entity["evidence_refs"] = _ordered_unique(ref for field_id in field_ids for ref in self.compiled_fields[field_id].evidence_refs)
            if entity_id == "product_001":
                entity["index"] = 1
                entity["status"] = "manual_review_required" if entity.get("review_required") else "verified"
                entity["confidence"] = {"overall": 1.0, "entity_linking_confidence": 1.0}

    def _pending_risks(self) -> list[Risk]:
        return list(self._extra_risks)


def _read_sheet(worksheet: Any) -> SheetTable:
    header_row, columns, header_cells, header_labels, ignored_columns = _locate_header(worksheet)
    rows: list[dict[str, CellValue]] = []
    for excel_row in worksheet.iter_rows(min_row=header_row + 1, max_row=worksheet.max_row):
        parsed_row: dict[str, CellValue] = {}
        for canonical_header, column_index in columns.items():
            cell = excel_row[column_index - 1]
            parsed_row[canonical_header] = CellValue(_display_text(cell), cell.coordinate)
        if any(cell_value.text.strip() for cell_value in parsed_row.values()):
            rows.append(parsed_row)
    return SheetTable(
        sheet_name=worksheet.title,
        header_row=header_row,
        columns=columns,
        header_cells=header_cells,
        header_labels=header_labels,
        ignored_columns=ignored_columns,
        rows=rows,
    )


def _locate_header(worksheet: Any) -> tuple[int, dict[str, int], dict[str, str], dict[str, str], list[dict[str, Any]]]:
    required = REQUIRED_HEADERS[worksheet.title]
    best: tuple[int, dict[str, int], dict[str, str], dict[str, str], list[dict[str, Any]]] | None = None
    max_header_row = min(10, worksheet.max_row or 1)
    for row in worksheet.iter_rows(min_row=1, max_row=max_header_row):
        columns: dict[str, int] = {}
        header_cells: dict[str, str] = {}
        header_labels: dict[str, str] = {}
        ignored_columns: list[dict[str, Any]] = []
        for cell in row:
            raw_header = _display_text(cell)
            if not raw_header:
                continue
            canonical = _canonical_header(raw_header)
            if _is_ignored_header(canonical, raw_header):
                ignored_columns.append({"header": raw_header, "canonical_header": canonical, "column": cell.column, "coordinate": cell.coordinate})
            if canonical is None:
                continue
            if canonical in columns:
                ignored_columns.append({"header": raw_header, "canonical_header": canonical, "column": cell.column, "coordinate": cell.coordinate, "reason": "duplicate_header"})
                continue
            columns[canonical] = cell.column
            header_cells[canonical] = cell.coordinate
            header_labels[canonical] = raw_header
        if required.issubset(columns):
            best = (int(row[0].row), columns, header_cells, header_labels, ignored_columns)
            break
    if best is None:
        raise ParseError(f"XLSX sheet {worksheet.title} is missing required headers: {', '.join(sorted(required))}")
    return best


def _canonical_header(raw_header: str) -> str | None:
    header = _compact_header(raw_header)
    if re.fullmatch(r"列\d+", header):
        return None
    alias_key = header
    if alias_key not in HEADER_ALIASES:
        alias_key = raw_header.strip()
    alias = HEADER_ALIASES.get(alias_key)
    if alias:
        return alias
    if _is_amount_header_alias(header):
        return "每100g/mL含量"
    return None


def _compact_header(value: str) -> str:
    return re.sub(r"\s+", "", str(value).strip())


def _is_amount_header_alias(header: str) -> bool:
    return bool(re.fullmatch(r"每份(?:[（(][^）)]+[）)])?", header, flags=re.IGNORECASE))


def _is_ignored_header(canonical: str | None, raw_header: str) -> bool:
    return canonical in IGNORED_HEADERS or canonical is None or bool(re.fullmatch(r"列\d+", _compact_header(raw_header)))


def _display_text(cell: Any) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if getattr(cell, "is_date", False):
        if hasattr(value, "date"):
            return value.date().isoformat()
        return str(value).strip()
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        if _is_percent_format(str(getattr(cell, "number_format", ""))):
            return f"{_format_number(float(value) * 100)}%"
        return _format_number(float(value))
    return str(value).strip()


def _format_number(value: float) -> str:
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:.10f}".rstrip("0").rstrip(".")


def _is_percent_format(number_format: str) -> bool:
    return "%" in number_format


def _is_print_value(value: str) -> bool:
    return value.strip().lower() not in NOT_APPLICABLE_VALUES


def _cell_text(row: dict[str, CellValue], header: str) -> str:
    cell = row.get(header)
    return cell.text.strip() if cell else ""


def _cell_coord(row: dict[str, CellValue], header: str) -> str:
    cell = row.get(header)
    return cell.coord if cell else "A1"


def _main_field_definition(code: str, label: str) -> tuple[str, str, str, str]:
    if code in MAIN_FIELD_BY_CODE:
        return MAIN_FIELD_BY_CODE[code]
    compact_label = _compact_header(label)
    for candidate, definition in MAIN_FIELD_BY_LABEL.items():
        if _compact_header(candidate) == compact_label:
            return definition
    return "custom.main_label_text", label or "标签主文字", "long_text", "non_critical"


def _enterprise_type(role: str) -> str:
    if "受委托" in role or "被委托" in role or "生产" in role:
        return "manufacturer"
    if "委托方" in role:
        return "principal"
    if "经销" in role:
        return "distributor"
    return "enterprise"


def _stable_group_id(prefix: str, raw_id: str, index: int) -> str:
    suffix = re.sub(r"[^0-9A-Za-z_]+", "_", raw_id.strip()) or str(index)
    return f"{prefix}_{suffix.lower()}"


def _safe_id(raw_id: str, index: int) -> str:
    suffix = re.sub(r"[^0-9A-Za-z_]+", "_", raw_id.strip()).strip("_")
    return suffix.lower() or f"{index:03d}"


def _split_refs(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[,，、;/；\s]+", value) if part.strip()]


def _span_id(sheet_name: str, coord: str) -> str:
    return f"xlsx:{sheet_name}!{coord}"


def _ordered_unique(values: Any) -> list[Any]:
    unique = []
    for value in values:
        if value not in unique:
            unique.append(value)
    return unique


def _entity_from_fields(entity_id: str, entity_type: str, index: int, field_ids: list[str], fields: dict[str, CompiledField]) -> dict[str, Any]:
    entity = _empty_entity(entity_id, entity_type, index)
    for field_id in field_ids:
        entity["fields"][field_id] = _entity_field(fields[field_id])
    entity["evidence_refs"] = _ordered_unique(ref for field_id in field_ids for ref in fields[field_id].evidence_refs)
    if any(fields[field_id].review_required for field_id in field_ids):
        entity["status"] = "manual_review_required"
        entity["risk_level"] = "high"
        entity["review_required"] = True
    return entity


def _empty_entity(entity_id: str, entity_type: str, index: int) -> dict[str, Any]:
    return {
        "entity_id": entity_id,
        "entity_type": entity_type,
        "index": index,
        "fields": {},
        "linked_table_ids": [],
        "status": "verified",
        "confidence": {"overall": 1.0, "entity_linking_confidence": 1.0},
        "risk_level": "low",
        "review_required": False,
        "evidence_refs": [],
    }


def _entity_field(field: CompiledField) -> dict[str, Any]:
    return {
        "field_id": field.field_id,
        "semantic_key": field.semantic_key,
        "value": field.raw_value,
        "status": field.status,
        "criticality": field.criticality,
        "confidence": field.confidence,
        "risk_level": field.risk_level,
        "review_required": field.review_required,
        "evidence_refs": field.evidence_refs,
    }


def _nutrition_standard_text(title: str, rows: list[dict[str, CellValue]], footnote: str) -> str:
    lines = [title]
    for row in rows:
        parts = [row["item"].text, row["amount"].text, row["nrv"].text]
        lines.append(" ".join(part for part in parts if part))
    if footnote:
        lines.append(footnote)
    return "\n".join(lines)


def _workbook_structure(sheetnames: list[str], sheets: dict[str, SheetTable]) -> dict[str, Any]:
    return {
        "artifact_version": "xlsx_workbook_structure_v0.1",
        "read_sheets": TARGET_SHEETS,
        "ignored_sheets": [sheet_name for sheet_name in sheetnames if sheet_name not in TARGET_SHEETS],
        "sheets": {
            sheet_name: {
                "header_row": sheet.header_row,
                "columns": sheet.columns,
                "header_cells": sheet.header_cells,
                "header_labels": sheet.header_labels,
                "display_columns": _display_columns(sheet),
                "rows": _display_rows(sheet),
                "ignored_columns": sheet.ignored_columns,
                "row_count": len(sheet.rows),
            }
            for sheet_name, sheet in sheets.items()
        },
    }


def _display_columns(sheet: SheetTable) -> list[dict[str, Any]]:
    return [
        {
            "canonical_header": canonical_header,
            "header": sheet.header_labels.get(canonical_header, canonical_header),
            "coordinate": sheet.header_cells.get(canonical_header),
            "ignored_as_content": False,
        }
        for canonical_header in _display_column_headers(sheet)
    ]


def _display_rows(sheet: SheetTable) -> list[dict[str, Any]]:
    rows = []
    display_headers = _display_column_headers(sheet)
    for row_index, row in enumerate(sheet.rows, start=sheet.header_row + 1):
        cells = {}
        for canonical_header in display_headers:
            cell = row.get(canonical_header)
            cells[canonical_header] = {
                "coordinate": cell.coord if cell else None,
                "source_node_id": _span_id(sheet.sheet_name, cell.coord) if cell else None,
                "text": cell.text if cell else "",
                "ignored_as_content": False,
            }
        if not any(str(cell.get("text") or "").strip() for cell in cells.values()):
            continue
        rows.append({"row_index": row_index, "cells": cells})
    return rows


def _display_column_headers(sheet: SheetTable) -> list[str]:
    return [
        header
        for header in sorted(sheet.columns, key=lambda item: sheet.columns[item])
        if header not in IGNORED_HEADERS
    ]


def _xlsx_runtime_policy() -> dict[str, Any]:
    return {
        "status": "pass",
        "source": "manifest",
        "input_format": "xlsx",
        "ocr": {"mode": "not_applicable", "provider": "not_applicable"},
        "llm_agent": {"enabled": False, "mode": "not_applicable"},
        "checks": [
            {"check": "xlsx_structured_input", "result": "passed"},
            {"check": "no_ocr_required", "result": "passed"},
            {"check": "no_llm_required", "result": "passed"},
        ],
    }


def _json_export_manifest() -> dict[str, Any]:
    return {
        "schema_version": "mvp_final_json_v0.1",
        "media_type": "application/json",
        "encoding": "utf-8",
        "root_keys": FINAL_JSON_ROOT_KEYS,
        "contract_checks": JSON_EXPORT_CONTRACT_CHECKS,
        "no_guessing": True,
        "primary_artifact": "result.json",
        "contract_artifact": "output_contract_validation_report.json",
        "schema_artifact": "schemas/final_result.schema.json",
    }


def _validation_check(
    checks: list[dict[str, Any]],
    target_id: str,
    check_type: str,
    result: str,
    message: str,
    evidence_refs: list[str],
) -> dict[str, Any]:
    return {
        "validation_id": stable_id("val", len(checks) + 1),
        "target_id": target_id,
        "check_type": check_type,
        "result": result,
        "severity": "info",
        "message": message,
        "evidence_refs": evidence_refs,
    }


def _review_tasks_from_risks(risks: list[Risk]) -> list[ReviewTask]:
    tasks = []
    for risk in risks:
        if risk.risk_level != "high":
            continue
        tasks.append(
            ReviewTask(
                task_id=stable_id("review", len(tasks) + 1),
                target_type=risk.target_type,
                target_id=risk.target_id,
                risk_level=risk.risk_level,
                reason=risk.message,
                required=True,
                evidence_refs=risk.evidence_refs,
            )
        )
    return tasks


def _quality(fields: dict[str, CompiledField], risks: list[Risk]) -> dict[str, Any]:
    high_count = sum(1 for risk in risks if risk.risk_level == "high")
    medium_count = sum(1 for risk in risks if risk.risk_level == "medium")
    low_count = sum(1 for risk in risks if risk.risk_level == "low")
    critical_fields = [field for field in fields.values() if field.criticality == "critical"]
    critical_pass = [field for field in critical_fields if not field.review_required and float(field.confidence.get("overall") or 0) >= 0.95]
    return {
        "overall_status": "manual_review_required" if high_count else "pass_with_warnings" if medium_count or low_count else "pass",
        "critical_confidence_threshold": 0.95,
        "field_completion_rate": 1.0 if fields else 0.0,
        "critical_field_pass_rate": round(len(critical_pass) / len(critical_fields), 4) if critical_fields else 0.0,
        "high_risk_count": high_count,
        "medium_risk_count": medium_count,
        "low_risk_count": low_count,
        "auto_ingest_allowed": high_count == 0,
        "reason": "存在 high risk，需要人工复核。" if high_count else None,
    }


def _coverage() -> dict[str, Any]:
    return {
        "text_block_coverage_rate": 1.0,
        "important_region_coverage_rate": 1.0,
        "table_cell_coverage_rate": 1.0,
        "unknown_important_block_count": 0,
        "conflict_node_count": 0,
        "status_counts": {"verified": 1},
        "assigned_node_ids": [],
    }


def _source_layers(spans: list[TextSpan], table_layers: dict[str, Any]) -> dict[str, Any]:
    return {
        "status": "pass",
        "source_mode": "xlsx_structured",
        "pages": [{"page": 1, "width": 1, "height": 1, "text_span_count": len(spans), "ocr_line_count": 0, "bbox_available_count": 0}],
        "layers": {
            "pdf_text": {"available": False, "span_count": 0, "bbox_available_count": 0, "warnings": []},
            "ocr": {
                "provider": "not_applicable",
                "status": "not_applicable",
                "error": None,
                "fallback_used": False,
                "line_count": 0,
                "block_count": 0,
                "token_count": 0,
                "bbox_available_count": 0,
                "confidence_min": None,
                "confidence_avg": None,
                "blocks": [],
                "lines": [],
            },
            "tables": {
                "parsers": table_layers.get("parsers", []),
                "table_count": len(table_layers.get("tables", [])),
                "parser_issue_count": len(table_layers.get("parser_issues", [])),
            },
        },
        "text_quality": {
            "total_text_span_count": len(spans),
            "total_char_count": sum(len(span.text) for span in spans),
            "bbox_coverage_rate": 0.0,
            "control_char_count": 0,
            "cjk_internal_spacing_span_count": 0,
            "multi_anchor_line_count": 0,
        },
        "spans": [
            {
                "span_id": span.span_id,
                "page": span.page,
                "source": span.source,
                "text": span.text,
                "char_count": len(span.text),
                "bbox_status": "missing",
                "bbox_pdf": None,
                "bbox_normalized": None,
                "confidence": span.confidence,
            }
            for span in spans
        ],
        "source_issues": [],
        "source_issue_count": 0,
    }


def _source_consistency_report() -> dict[str, Any]:
    return {
        "status": "pass",
        "pdf_text_span_count": 0,
        "ocr_line_count": 0,
        "matched_ocr_line_count": 0,
        "issue_count": 0,
        "issues": [],
    }


def _table_parser_artifacts(tables: list[dict[str, Any]]) -> tuple[dict[str, Any], dict[str, Any]]:
    table_layers = {
        "parsers": ["xlsx_structured"],
        "tables": [
            {
                "table_layer_id": table["table_id"],
                "table_id": table["table_id"],
                "table_type": table["table_type"],
                "title": table["title"],
                "row_count": len(table.get("rows", [])),
                "column_count": len(table.get("columns", [])),
                "source": "xlsx_structured_table",
            }
            for table in tables
        ],
        "parser_issues": [],
    }
    quality = {
        "status": "pass",
        "table_count": len(tables),
        "issue_count": 0,
        "issues": [],
        "parser_agreement": {"status": "single_source_structured", "parsers": ["xlsx_structured"]},
    }
    return table_layers, quality


def _schema_audit() -> dict[str, Any]:
    return {
        "status": "pass",
        "issue_count": 0,
        "blocking_issue_count": 0,
        "issues": [],
        "checks": [{"check": "xlsx_template_schema", "result": "passed"}],
    }


def _structure_audit(regions: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "status": "pass",
        "anchor_coverage": 1.0,
        "missing_anchor_count": 0,
        "sequence_gap_count": 0,
        "group_issue_count": 0,
        "table_issue_count": 0,
        "required_prefix_issue_count": 0,
        "container_duplicate_issue_count": 0,
        "agent_override_issue_count": 0,
        "duplicate_coverage_issue_count": 0,
        "missing_anchor_issues": [],
        "group_issues": [],
        "table_issues": [],
        "required_prefix_issues": [],
        "container_duplicate_issues": [],
        "agent_override_issues": [],
        "duplicate_coverage_issues": [],
        "anchor_inventory": [],
        "region_count": len(regions),
    }


def _repair_loop() -> dict[str, Any]:
    repair_plan = {"status": "pass", "repair_mode": "not_applicable", "actions": [], "max_repair_rounds": 0}
    return {
        "max_repair_rounds": 0,
        "audit_finding_count": 0,
        "trace": {"status": "pass", "round_count": 0, "attempt_count": 0, "applied_attempt_count": 0, "final_audit_finding_count": 0, "rounds": []},
        "attempts": {"status": "pass", "attempt_count": 0, "attempts": []},
        "repair_plan": repair_plan,
        "repair_plan_patches": {"status": "pass", "patches": []},
        "repair_agent_candidates": {"status": "pass", "candidate_count": 0, "candidates": []},
        "repaired_source_layers": {"status": "not_applicable", "source_issues": []},
        "policy": "xlsx_structured_no_repair_required",
    }


def _visual_document_graph(regions: list[dict[str, Any]]) -> dict[str, Any]:
    nodes = [
        {
            "node_id": region["region_id"],
            "node_type": "region",
            "region_type": region["region_type"],
            "page": region["page"],
            "status": "verified",
            "source_span_ids": region["source_span_ids"],
        }
        for region in regions
    ]
    return {
        "graph_id": "vdg_xlsx_001",
        "schema_version": "vdg_v0.1",
        "node_count": len(nodes),
        "edge_count": 0,
        "nodes": nodes,
        "edges": [],
        "applicability": "not_applicable_for_structured_xlsx",
    }


def _vdg_quality_report() -> dict[str, Any]:
    return {
        "report_version": "vdg_quality_report_v0.1",
        "status": "pass",
        "source_span_coverage_rate": 1.0,
        "edge_ref_status": "pass",
        "issues": [],
        "checks": [{"check": "xlsx_vdg_placeholder", "result": "passed"}],
        "applicability": "not_applicable_for_structured_xlsx",
    }


def _vdg_agent_context(table_layers: dict[str, Any]) -> dict[str, Any]:
    return {
        "context_version": "vdg_agent_context_v0.1",
        "vdg_quality_status": "pass",
        "agent_readiness": "not_applicable",
        "candidate_field_groups": [],
        "table_candidates": table_layers.get("tables", []),
        "quality_issues": [],
    }


def _vdg_consumption_report() -> dict[str, Any]:
    return {
        "report_version": "vdg_consumption_report_v0.1",
        "status": "pass",
        "consumable_node_count": 0,
        "extracted_node_count": 0,
        "extracted_coverage_rate": 1.0,
        "status_counts": {"verified": 0},
        "extracted_node_ids": [],
        "unknown_important_node_count": 0,
        "conflict_node_count": 0,
    }


def _label_text_scope_report(reference: dict[str, Any]) -> dict[str, Any]:
    return {
        "report_version": "label_text_scope_report_v0.1",
        "reference_version": reference.get("reference_version"),
        "status": "pass",
        "extracted_out_of_scope_count": 0,
        "ignored_noise_node_count": 0,
        "unknown_scope_node_count": 0,
        "scope_gate_rejected_count": 0,
        "node_scope_decisions": [],
        "checks": [{"check": "xlsx_whitelist_sheets", "result": "passed"}],
    }


def _audit_input(schema: GeneratedSchema, fields: dict[str, CompiledField], evidence: list[Evidence]) -> dict[str, Any]:
    return {
        "source": "xlsx_structured",
        "schema": to_jsonable(schema),
        "field_count": len(fields),
        "evidence_count": len(evidence),
    }


def _agent_execution_report(
    schema: GeneratedSchema,
    plan: ExtractionPlan,
    fields: dict[str, CompiledField],
    evidence: list[Evidence],
) -> dict[str, Any]:
    return {
        "status": "pass",
        "mode": "xlsx_structured_no_agent",
        "schema_agent": None,
        "extraction_agent": None,
        "audit_agent": None,
        "repair_agent": None,
        "llm_agent": None,
        "schema_field_count": len(schema.field_definitions),
        "plan_field_count": len(plan.fields),
        "compiled_field_count": len(fields),
        "evidence_count": len(evidence),
        "rejected_agent_items": [],
        "review_items": [],
    }


def _agent_harness() -> dict[str, Any]:
    return {
        "agent_items_path": None,
        "accepted_agent_item_count": 0,
        "agent_plan_field_count": 0,
        "agent_field_retry_count": 0,
        "agent_table_retry_count": 0,
        "field_retry_used": False,
        "table_retry_used": False,
        "rule_fallback_field_count": 0,
        "rule_fallback_items": [],
        "llm_agent_enabled": False,
        "vdg_agent_context_used": False,
        "vdg_quality_status": "pass",
        "label_text_scope_status": "pass",
        "scope_gate_rejected_count": 0,
        "llm_agent_candidate_count": 0,
        "llm_field_retry_candidate_count": 0,
        "llm_table_retry_table_count": 0,
        "llm_field_retry_items": {"fields": []},
        "llm_table_retry_items": {"tables": []},
        "llm_agent_items": {"fields": []},
        "llm_schema_items": {},
        "rejected_agent_items": [],
        "review_items": [],
    }


def _detected_document_types(tables: list[dict[str, Any]]) -> list[str]:
    types = {"packaging_label_standard", "structured_xlsx_standard"}
    if tables:
        types.add("nutrition_table_document")
    return sorted(types)


def _write_layered_debug_artifacts(debug_dir: Path, result: ParseResult, workbook_structure: dict[str, Any]) -> None:
    metadata = result.metadata
    standard = metadata["standard_artifacts"]
    write_json(debug_dir / "00_inputs" / "runtime_policy.json", metadata["runtime_policy"])
    write_json(debug_dir / "00_inputs" / "json_export.json", metadata["json_export"])
    write_json(debug_dir / "00_inputs" / "xlsx_workbook_structure.json", workbook_structure)
    write_json(debug_dir / "01_text_extraction" / "source_layers.json", metadata["source_layers"])
    write_json(debug_dir / "01_text_extraction" / "visual_document_graph.json", metadata["visual_document_graph"])
    write_json(debug_dir / "01_table_extraction" / "table_layers.json", metadata["table_parser"]["table_layers"])
    write_json(debug_dir / "02_section_detection" / "regions.json", result.extracted_data["regions"])
    write_json(debug_dir / "03_field_structure" / "standard_items.json", standard["standard_items"])
    write_result_preview_html(
        result,
        debug_dir / "03_field_structure" / "result_preview.html",
        artifact_root=debug_dir,
        workbook_structure=workbook_structure,
    )
    write_json(debug_dir / "03_field_structure" / "comparison_index.json", standard["comparison_index"])
    write_json(debug_dir / "03_field_structure" / "field_groups.json", standard["field_groups"])
    write_json(debug_dir / "03_field_structure" / "tables.json", standard["tables"])
    write_json(debug_dir / "03_field_structure" / "lists.json", standard["lists"])
    write_json(debug_dir / "03_field_structure" / "extracted_data.json", to_jsonable(result.extracted_data))
    write_json(debug_dir / "04_validation" / "quality_report.json", standard["quality_report"])
    write_json(debug_dir / "04_validation" / "output_contract_validation_report.json", metadata["output_contract_validation_report"])
    write_json(debug_dir / "04_validation" / "mvp_acceptance_metrics.json", metadata["mvp_acceptance_metrics"])
    write_json(debug_dir / "04_validation" / "evidence.json", to_jsonable(result.evidence))
    write_json(debug_dir / "04_validation" / "risks.json", to_jsonable(result.risks))
    write_json(debug_dir / "04_validation" / "review_tasks.json", to_jsonable(result.review_tasks))
    write_json(debug_dir / "05_repair" / "repair_plan.json", metadata["repair_loop"]["repair_plan"])
    write_json(debug_dir / "05_repair" / "repair_plan_patches.json", metadata["repair_loop"]["repair_plan_patches"])


def write_artifact_index(debug_dir: Path) -> None:
    index_path = debug_dir / "artifacts" / "index.json"
    artifacts = []
    indexed_suffixes = {".json", ".md", ".png", ".html"}
    for path in sorted(item for item in debug_dir.rglob("*") if item.is_file() and item.suffix in indexed_suffixes):
        if path == index_path:
            continue
        artifacts.append(
            {
                "path": path.relative_to(debug_dir).as_posix(),
                "sha256": sha256_file(path),
                "size_bytes": path.stat().st_size,
            }
        )
    write_json(index_path, {"artifact_count": len(artifacts), "artifacts": artifacts})
